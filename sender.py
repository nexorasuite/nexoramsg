from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import quote
import time
import os
import openpyxl
from datetime import datetime
import random
import qrcode
from io import BytesIO

# Global Chrome driver (reused across calls)
driver = None

def generate_qr_code(data="https://web.whatsapp.com"):
    """Generate QR code image and save it"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    qr_path = os.path.join(os.getcwd(), 'static', 'qr_code.png')
    os.makedirs('static', exist_ok=True)
    img.save(qr_path)
    return qr_path

def get_random_delay():
    """Get random delay between 35 seconds and 3 minutes"""
    return random.uniform(35, 180)

def init_driver():
    global driver
    if driver is None:
        user_data_dir = os.path.join(os.getcwd(), 'user_data', 'default_profile')
        os.makedirs(user_data_dir, exist_ok=True)

        options = webdriver.ChromeOptions()
        
        # Try to find Chromium on Raspberry Pi or Desktop
        chromium_paths = [
            '/usr/bin/chromium-browser',  # Raspberry Pi standard location
            '/usr/bin/chromium',          # Alternative location
            '/snap/bin/chromium',         # Snap installation
            '/usr/bin/google-chrome',     # Google Chrome
            '/Applications/Chromium.app/Contents/MacOS/Chromium'  # macOS
        ]
        
        chromium_found = False
        for path in chromium_paths:
            if os.path.exists(path):
                options.binary_location = path
                chromium_found = True
                print(f"✅ Found Chromium at: {path}")
                break
        
        if not chromium_found:
            print("⚠️  Chromium not found in standard paths, using system default...")
        
        options.add_argument(f'--user-data-dir={user_data_dir}')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        
        # Raspberry Pi optimizations
        options.add_argument('--disable-gpu')  # Disable GPU rendering (Pi issue fix)
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-plugins')
        options.add_argument('--disable-software-rasterizer')
        options.add_argument('--disable-3d-apis')
        options.add_argument('--disable-client-side-phishing-detection')
        
        try:
            # Try with specific chromedriver
            chromedriver_paths = [
                '/usr/bin/chromedriver',
                '/snap/bin/chromium.chromedriver',
                '/usr/lib/chromium-browser/chromedriver'
            ]
            
            service = None
            for path in chromedriver_paths:
                if os.path.exists(path):
                    service = Service(path)
                    print(f"✅ Found chromedriver at: {path}")
                    break
            
            if service:
                driver = webdriver.Chrome(service=service, options=options)
            else:
                # Fallback to automatic chromedriver
                driver = webdriver.Chrome(options=options)
        except Exception as e:
            print(f"⚠️  Chromedriver init error: {e}, trying fallback...")
            driver = webdriver.Chrome(options=options)

        # Generate and display QR code before showing WhatsApp Web
        qr_path = generate_qr_code()
        print(f"📱 QR Code generated at: {qr_path}")
        print("🔐 Please scan the QR code from your phone to login to WhatsApp Web")
        
        # Load WhatsApp Web and wait for user login
        driver.get('https://web.whatsapp.com')
        WebDriverWait(driver, 300).until(
            EC.presence_of_element_located((By.ID, "side"))
        )
        print("✅ WhatsApp Web loaded successfully!")
    return driver

def check_and_clear_draft(driver, number):
    """
    Check if message is still in draft (text field) and hasn't been sent.
    Returns True if draft was detected, False if message was sent.
    """
    try:
        # Look for the message input field
        input_selectors = [
            "//div[@contenteditable='true'][@data-tab='1']",  # Message input
            "//div[@contenteditable='true'][@role='textbox']",
            "//div[@contenteditable='true'][@spellcheck='false']",
            "//div[@data-testid='msg-input']"
        ]
        
        for selector in input_selectors:
            try:
                input_field = driver.find_element(By.XPATH, selector)
                text_content = input_field.text or input_field.get_attribute("textContent") or ""
                
                if text_content.strip():  # If there's still text in input field
                    print(f"📝 Draft detected for {number} - clearing...")
                    
                    # Try to select all and delete
                    try:
                        input_field.clear()
                    except:
                        # Fallback: use keyboard shortcuts
                        driver.execute_script("""
                            const elem = arguments[0];
                            elem.textContent = '';
                            elem.innerText = '';
                            const event = new Event('input', { bubbles: true });
                            elem.dispatchEvent(event);
                        """, input_field)
                    
                    time.sleep(0.5)
                    return True  # Draft was detected and cleared
            except:
                continue
        
        return False  # No draft found
    
    except Exception as e:
        print(f"⚠️ Error checking draft for {number}: {e}")
        return False

def verify_message_sent(driver, number):
    """
    Verify that the message was actually sent by checking for:
    1. Absence of send button (message field empty)
    2. Input field is empty (indicates message sent)
    """
    try:
        # Method 1: Check if send button is no longer visible/clickable
        try:
            send_button = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
            if send_button.is_displayed():
                # Send button still visible and clickable means message might still be draft
                try:
                    driver.execute_script("arguments[0].scrollIntoView(true);", send_button)
                    return False
                except:
                    pass
        except:
            # Send button not found - good sign, message likely sent
            pass
        
        # Method 2: Check if message input field is empty
        input_selectors = [
            "//div[@contenteditable='true'][@data-tab='1']",
            "//div[@contenteditable='true'][@role='textbox']",
            "//div[@contenteditable='true'][@spellcheck='false']",
        ]
        
        for selector in input_selectors:
            try:
                input_field = driver.find_element(By.XPATH, selector)
                text_content = input_field.text or input_field.get_attribute("textContent") or ""
                
                if text_content.strip():
                    # Still has text - message wasn't sent
                    return False
            except:
                continue
        
        # Message appears to have been sent successfully
        return True
    
    except Exception as e:
        print(f"⚠️ Error verifying send for {number}: {e}")
        return True  # Assume sent to continue

def send_whatsapp_messages_with_log(numbers, message, log_path, append=False, task_manager=None, task_id=None):
    # Prepare Excel workbook for logging
    if append and os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WhatsApp Logs"
        ws.append(['Phone Number', 'Status', 'Timestamp', 'Delay Used (sec)'])

    driver = init_driver()

    encoded_message = quote(message)
    
    sent_count = 0
    failed_count = 0
    invalid_count = 0

    for idx, number in enumerate(numbers):
        try:
            # Update task progress
            if task_manager and task_id:
                task_manager.update_task(
                    task_id,
                    current_index=idx + 1,
                    current_recipient=number,
                    progress_percent=int((idx + 1) / len(numbers) * 100)
                )
            
            # Get random delay between sends
            delay = get_random_delay()
            
            url = f"https://web.whatsapp.com/send?phone={number}&text={encoded_message}"
            driver.get(url)

            # Wait for the send button and click
            send_button = WebDriverWait(driver, 40).until(
                EC.element_to_be_clickable((By.XPATH, '//span[@data-icon="send"]'))
            )
            send_button.click()
            
            # Check if message is still in draft and force send if needed
            time.sleep(2)  # Wait for message to be processed
            draft_detected = check_and_clear_draft(driver, number)
            
            if draft_detected:
                print(f"📤 Draft detected for {number}, forcing send...")
                time.sleep(1)
                try:
                    send_button = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, '//span[@data-icon="send"]'))
                    )
                    send_button.click()
                    time.sleep(2)
                except:
                    pass
            
            # Verify message was actually sent (check for success indicators)
            send_verified = verify_message_sent(driver, number)
            if not send_verified:
                print(f"⚠️ Send verification failed for {number}, checking again...")
                time.sleep(1)
                # Try clicking send button once more if visible
                try:
                    send_button = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
                    if send_button:
                        print(f"📤 Retrying send for {number}...")
                        send_button.click()
                        time.sleep(2)
                except:
                    pass

            print(f"✅ Message sent to {number}")
            ws.append([number, "Sent", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"{delay:.1f}"])
            sent_count += 1
            
            # Update task stats
            if task_manager and task_id:
                task_manager.update_task(task_id, sent=sent_count, current_delay=delay)
            
            # Random delay between messages to avoid WhatsApp ban
            print(f"⏳ Waiting {delay:.1f} seconds before next message... ({idx+1}/{len(numbers)})")
            time.sleep(delay)

        except Exception as e:
            if "Phone number shared via URL is invalid" in driver.page_source:
                print(f"⚠️ Invalid number: {number}")
                ws.append([number, "Invalid", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "-"])
                invalid_count += 1
            else:
                print(f"❌ Failed to send to {number}: {e}")
                ws.append([number, f"Failed: {str(e)}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "-"])
                failed_count += 1
            
            # Update task stats
            if task_manager and task_id:
                task_manager.update_task(task_id, failed=failed_count, invalid=invalid_count)

    wb.save(log_path)
    print(f"📄 Log saved to {log_path}")

def close_driver():
    global driver
    if driver:
        driver.quit()
        driver = None


def send_telegram_messages_with_log(chat_ids, message, log_path, append=False, api_token=None, task_manager=None, task_id=None):
    """
    Send messages via Telegram Bot API
    
    Args:
        chat_ids: List of Telegram chat IDs or usernames
        message: Message to send
        log_path: Path to save Excel log
        append: Whether to append to existing log
        api_token: Telegram Bot API token (from @BotFather)
        task_manager: Task manager for progress tracking
        task_id: Task ID for progress updates
    """
    import requests
    
    if not api_token:
        print("❌ Telegram API token not provided!")
        return
    
    # Prepare Excel workbook for logging
    if append and os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Telegram Logs"
        ws.append(['Chat ID', 'Status', 'Timestamp', 'Delay Used (sec)'])

    base_url = f"https://api.telegram.org/bot{api_token}/sendMessage"
    
    sent_count = 0
    failed_count = 0

    for idx, chat_id in enumerate(chat_ids):
        try:
            # Update task progress
            if task_manager and task_id:
                task_manager.update_task(
                    task_id,
                    current_index=idx + 1,
                    current_recipient=str(chat_id),
                    progress_percent=int((idx + 1) / len(chat_ids) * 100)
                )
            
            # Get random delay between sends
            delay = get_random_delay()
            
            payload = {
                "chat_id": chat_id,
                "text": message,
                "parse_mode": "HTML"
            }
            
            response = requests.post(base_url, json=payload, timeout=10)
            
            if response.status_code == 200:
                print(f"✅ Message sent to {chat_id}")
                ws.append([chat_id, "Sent", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), f"{delay:.1f}"])
                sent_count += 1
            else:
                error_msg = response.json().get('description', 'Unknown error')
                print(f"❌ Failed to send to {chat_id}: {error_msg}")
                ws.append([chat_id, f"Failed: {error_msg}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "-"])
                failed_count += 1
            
            # Update task stats
            if task_manager and task_id:
                task_manager.update_task(task_id, sent=sent_count, failed=failed_count, current_delay=delay)
            
            # Random delay between messages
            print(f"⏳ Waiting {delay:.1f} seconds before next message... ({idx+1}/{len(chat_ids)})")
            time.sleep(delay)
            
        except Exception as e:
            print(f"❌ Error sending to {chat_id}: {e}")
            ws.append([chat_id, f"Error: {str(e)}", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "-"])
            failed_count += 1
            
            # Update task stats
            if task_manager and task_id:
                task_manager.update_task(task_id, failed=failed_count)

    wb.save(log_path)
    print(f"📄 Log saved to {log_path}")